import sys
import random
import gurobipy as gp
from gurobipy import GRB
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QSpinBox, QPushButton,
    QFormLayout, QInputDialog, QMessageBox, QTabWidget, QGraphicsView,
    QGraphicsScene, QGraphicsRectItem, QScrollArea, QGroupBox,
    QHBoxLayout, QTableWidget, QTableWidgetItem, QCheckBox
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QBrush, QPen, QColor
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# ------------------------
# DEFAULT DATA
# ------------------------
DEFAULT_TRUCKS = {
    "T1": {"L":70,"W":45,"H":40,"max_weight":260},
    "T2": {"L":70,"W":45,"H":40,"max_weight":460},
    "T3": {"L":60,"W":40,"H":35,"max_weight":840},
}

DEFAULT_ITEMS = {
    "A": {"L":5,"W":10,"H":5,"weight":25},
    "B": {"L":10,"W":10,"H":10,"weight":10},
    "C": {"L":4,"W":5,"H":3,"weight":6},
}

EXCEL_FILE = "packing_data.xlsx"

# ------------------------
# Random colors for items
# ------------------------
ITEM_COLORS = {}
def get_item_color(item):
    if item not in ITEM_COLORS:
        ITEM_COLORS[item] = (random.random(), random.random(), random.random(), 0.8)
    return ITEM_COLORS[item]

# ------------------------
# Excel functions
# ------------------------
def save_to_excel(items, trucks, filename=EXCEL_FILE):
    df_items = pd.DataFrame.from_dict(items, orient='index')
    df_items.index.name = 'Item'
    df_trucks = pd.DataFrame.from_dict(trucks, orient='index')
    df_trucks.index.name = 'Truck'
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_items.to_excel(writer, sheet_name="Items")
        df_trucks.to_excel(writer, sheet_name="Trucks")

def load_from_excel(filename=EXCEL_FILE):
    if not Path(filename).exists():
        return DEFAULT_ITEMS.copy(), DEFAULT_TRUCKS.copy()
    xls = pd.ExcelFile(filename)
    df_items = pd.read_excel(xls, sheet_name="Items", index_col=0)
    df_trucks = pd.read_excel(xls, sheet_name="Trucks", index_col=0)

    items = df_items.fillna(0).to_dict(orient="index")
    for k,v in items.items():
        for key,value in v.items():
            if pd.isna(value):
                items[k][key] = 0
            else:
                items[k][key] = int(value) if float(value) == int(float(value)) else float(value)
    
    trucks = df_trucks.fillna(0).to_dict(orient="index")
    for k,v in trucks.items():
        for key,value in v.items():
            if pd.isna(value):
                trucks[k][key] = 0
            else:
                trucks[k][key] = int(value) if float(value) == int(float(value)) else float(value)
    
    return items, trucks

def save_solution_to_excel(solution, truck_weights, items, trucks, filename=EXCEL_FILE):
    rows = []
    for t, item_list in solution.items():
        if not item_list:
            continue
        counts = {}
        for it in item_list:
            counts[it["item"]] = counts.get(it["item"], 0) + 1
        used_weight = truck_weights[t]
        max_weight = trucks[t]["max_weight"]
        fullness = round(100 * used_weight / max_weight, 2) if max_weight>0 else 0
        item_types = list(counts.keys())
        for idx, item_type in enumerate(item_types):
            rows.append({
                "Truck": t if idx==0 else "-",
                "Weight": f"{used_weight} / {max_weight}" if idx==0 else "-",
                "Item": item_type,
                "Count": counts[item_type],
                "Item Weight": items[item_type]["weight"],
                "Fullness %": f"{fullness}" if idx==0 else "-"
            })
    df_solution = pd.DataFrame(rows)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    with pd.ExcelWriter(filename, engine="openpyxl", mode='a', if_sheet_exists="replace") as writer:
        df_solution.to_excel(writer, sheet_name=f"Solution_{timestamp}", index=False)

def delete_all_history():
    if not Path(EXCEL_FILE).exists():
        return
    wb = load_workbook(EXCEL_FILE)
    sheets_to_remove = [s for s in wb.sheetnames if s.startswith("Solution_")]
    for sheet in sheets_to_remove:
        wb.remove(wb[sheet])
    wb.save(EXCEL_FILE)

# ------------------------
# Gurobi Optimization
# ------------------------
def optimize_packing_with_items(items, item_qty, trucks):
    model = gp.Model("Packing")
    x = {}
    y = {}
    for t in trucks:
        y[t] = model.addVar(vtype=GRB.BINARY)
    for i in items:
        for j in range(item_qty[i]):
            for t in trucks:
                x[i,j,t] = model.addVar(vtype=GRB.BINARY)
    for i in items:
        for j in range(item_qty[i]):
            model.addConstr(gp.quicksum(x[i,j,t] for t in trucks)==1)
    for t in trucks:
        model.addConstr(gp.quicksum(x[i,j,t]*items[i]["weight"] for i in items for j in range(item_qty[i])) <= trucks[t]["max_weight"]*y[t])
    for i in items:
        l,w,h = items[i]["L"], items[i]["W"], items[i]["H"]
        for j in range(item_qty[i]):
            for t in trucks:
                if l>trucks[t]["L"] or w>trucks[t]["W"] or h>trucks[t]["H"]:
                    model.addConstr(x[i,j,t]==0)
    model.setObjective(gp.quicksum(y[t] for t in trucks), GRB.MINIMIZE)
    model.optimize()
    if model.status != GRB.OPTIMAL and model.status != GRB.SUBOPTIMAL:
        QMessageBox.warning(None,"Optimization Error","No feasible solution found!")
        return {t: [] for t in trucks}, {t:0 for t in trucks}
    solution = {t: [] for t in trucks}
    truck_weights = {t:0 for t in trucks}
    for i in items:
        for j in range(item_qty[i]):
            for t in trucks:
                var = x.get((i,j,t))
                if var and var.X>0.5:
                    solution[t].append({"item":i,"weight":items[i]["weight"],
                                        "L":items[i]["L"],"W":items[i]["W"],"H":items[i]["H"]})
                    truck_weights[t]+=items[i]["weight"]
    return solution, truck_weights

# ------------------------
# 2D Visualization
# ------------------------
class Packing2DWindow(QMainWindow):
    def __init__(self, solution, trucks):
        super().__init__()
        self.setWindowTitle("2D Packing Visualization")
        self.resize(900,700)
        self.solution = solution
        self.trucks = trucks
        self.initUI()
    
    def initUI(self):
        view = QGraphicsView()
        scene = QGraphicsScene()
        view.setScene(scene)
        self.setCentralWidget(view)

        x_offset = 20
        y_offset = 20
        spacing = 50

        for t_name, items_list in self.solution.items():
            truck = self.trucks[t_name]
            truck_rect = QGraphicsRectItem(x_offset, y_offset, truck["L"]*3, truck["W"]*3)
            truck_rect.setPen(QPen(Qt.black, 2))
            truck_rect.setBrush(QBrush(Qt.transparent))
            scene.addItem(truck_rect)

            x_item = x_offset + 1
            y_item = y_offset + 1
            max_row_height = 0
            layer_height_offset = 0

            items_sorted = sorted(items_list, key=lambda it: -it["weight"])

            for it in items_sorted:
                L = it["L"]*3
                W = it["W"]*3

                if x_item + L > x_offset + truck["L"]*3:
                    x_item = x_offset + 1
                    y_item += max_row_height + 1
                    max_row_height = 0

                if y_item + W > y_offset + truck["W"]*3:
                    layer_height_offset += max_row_height + 5
                    y_item = y_offset + 1
                    x_item = x_offset + 1
                    max_row_height = 0

                color = QColor.fromRgbF(*get_item_color(it["item"]))
                rect = QGraphicsRectItem(x_item, y_item + layer_height_offset, L, W)
                rect.setBrush(QBrush(color))
                rect.setPen(QPen(Qt.black))
                scene.addItem(rect)

                x_item += L + 1
                max_row_height = max(max_row_height, W)

            y_offset += truck["W"]*3 + spacing + layer_height_offset

# ------------------------
# Main GUI with multi-step workflow
# ------------------------
class PackingWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Truck Packing Optimization Tool")
        self.resize(1400,900)
        self.items, self.trucks = load_from_excel()
        self.solution, self.truck_weights = None, None
        self.selected_items = []
        self.selected_trucks = []
        self.initUI()

    def initUI(self):
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # ------------------------
        # Step 1: Item Selection
        # ------------------------
        self.item_tab = QWidget()
        item_layout = QVBoxLayout()
        self.item_tab.setLayout(item_layout)
        self.tabs.addTab(self.item_tab,"Step 1: Select Items")

        self.item_checkboxes = {}
        for i in self.items:
            cb = QCheckBox(f"{i} (Length:{self.items[i]['L']} Width:{self.items[i]['W']} Height:{self.items[i]['H']} Weight:{self.items[i]['weight']})")
            self.item_checkboxes[i]=cb
            item_layout.addWidget(cb)

        # Item action buttons
        item_btn_layout = QHBoxLayout()
        for text, func, color in [("Add", self.add_item_dialog, "#4CAF50"),
                                  ("Edit", self.modify_item_dialog, "#FFA500"),
                                  ("Delete", self.delete_item_dialog, "#f44336")]:
            b = QPushButton(text)
            b.clicked.connect(func)
            b.setStyleSheet(f"background-color:{color}; color:white; border-radius:5px; padding:5px;")
            item_btn_layout.addWidget(b)
        item_layout.addLayout(item_btn_layout)

        # Next button
        next_item_btn = QPushButton("Next: Select Trucks")
        next_item_btn.setStyleSheet("background-color:#2196F3; color:white; border-radius:5px; padding:5px;")
        next_item_btn.clicked.connect(self.goto_truck_tab)
        item_layout.addWidget(next_item_btn)

        # ------------------------
        # Step 2: Truck Selection
        # ------------------------
        self.truck_tab = QWidget()
        truck_layout = QVBoxLayout()
        self.truck_tab.setLayout(truck_layout)
        self.tabs.addTab(self.truck_tab,"Step 2: Select Trucks")

        self.truck_checkboxes = {}
        for t in self.trucks:
            cb = QCheckBox(f"{t} (L:{self.trucks[t]['L']} W:{self.trucks[t]['W']} H:{self.trucks[t]['H']} MaxWt:{self.trucks[t]['max_weight']})")
            self.truck_checkboxes[t]=cb
            truck_layout.addWidget(cb)

        # Truck action buttons
        truck_btn_layout = QHBoxLayout()
        for text, func, color in [("Add", self.add_truck_dialog, "#4CAF50"),
                                  ("Edit", self.modify_truck_dialog, "#FFA500"),
                                  ("Delete", self.delete_truck_dialog, "#f44336")]:
            b = QPushButton(text)
            b.clicked.connect(func)
            b.setStyleSheet(f"background-color:{color}; color:white; border-radius:5px; padding:5px;")
            truck_btn_layout.addWidget(b)
        truck_layout.addLayout(truck_btn_layout)

        # Next button
        next_truck_btn = QPushButton("Next: Optimization")
        next_truck_btn.setStyleSheet("background-color:#2196F3; color:white; border-radius:5px; padding:5px;")
        next_truck_btn.clicked.connect(self.goto_main_tab)
        truck_layout.addWidget(next_truck_btn)

        # ------------------------
        # Step 3: Main Optimization Tab
        # ------------------------
        self.main_tab = QWidget()
        self.main_layout = QVBoxLayout()
        self.main_tab.setLayout(self.main_layout)
        self.tabs.addTab(self.main_tab,"Step 3: Optimization")

        # Form layout for selected items
        self.form_layout = QFormLayout()
        self.spin_boxes = {}
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        form_widget = QWidget()
        form_widget.setLayout(self.form_layout)
        scroll.setWidget(form_widget)
        self.main_layout.addWidget(scroll,2)

        # Optimization buttons
        opt_group = QGroupBox("Optimization Actions")
        opt_layout = QHBoxLayout()
        for text, func, color in [("Run Optimization", self.run_optimization, "#2196F3"),
                                  ("Save Solution", self.save_solution_dialog, "#9C27B0"),
                                  ("Show 2D Solution", self.show_2d, "#FF9800")]:
            b = QPushButton(text)
            b.clicked.connect(func)
            b.setStyleSheet(f"background-color:{color}; color:white; border-radius:5px; padding:5px;")
            opt_layout.addWidget(b)
        opt_group.setLayout(opt_layout)
        self.main_layout.addWidget(opt_group)

        # Results table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Truck","Item","Count","Weight","Fullness %"])
        self.main_layout.addWidget(self.table,3)

        # History tab
        self.history_tab = QWidget()
        self.history_layout = QVBoxLayout()
        self.history_tab.setLayout(self.history_layout)
        self.tabs.addTab(self.history_tab,"History")

        self.history_table = QTableWidget()
        self.history_layout.addWidget(self.history_table)

        self.delete_history_btn = QPushButton("Delete All History")
        self.delete_history_btn.setStyleSheet("background-color:#f44336; color:white; border-radius:5px; padding:5px;")
        self.delete_history_btn.clicked.connect(self.delete_history)
        self.history_layout.addWidget(self.delete_history_btn)

        self.load_history()

    # ------------------------
    # Step transitions
    # ------------------------
    def goto_truck_tab(self):
        self.selected_items = [i for i,cb in self.item_checkboxes.items() if cb.isChecked()]
        if not self.selected_items:
            QMessageBox.warning(self,"Warning","Select at least one item"); return
        self.tabs.setCurrentWidget(self.truck_tab)

    def goto_main_tab(self):
        self.selected_trucks = [t for t,cb in self.truck_checkboxes.items() if cb.isChecked()]
        if not self.selected_trucks:
            QMessageBox.warning(self,"Warning","Select at least one truck"); return
        self.update_main_form()
        self.tabs.setCurrentWidget(self.main_tab)

    # ------------------------
    # Main tab form update
    # ------------------------
    def update_main_form(self):
        while self.form_layout.rowCount()>0:
            self.form_layout.removeRow(0)
        self.spin_boxes.clear()
        for i in self.selected_items:
            spin = QSpinBox()
            spin.setRange(0,1000)
            spin.setValue(1)
            self.spin_boxes[i]=spin
            self.form_layout.addRow(f"Quantity of {i}:", spin)

    # ------------------------
    # Item / Truck dialogs
    # ------------------------
    def add_item_dialog(self):
        name,ok=QInputDialog.getText(self,"New Item","Enter item name:")
        if not ok or not name.strip(): return
        name=name.strip()
        if name in self.items:
            QMessageBox.warning(self,"Error",f"Item {name} exists"); return
        params={}
        for param in ["Length","Weight","Height","weight"]:
            while True:
                value,ok=QInputDialog.getDouble(self,"New Item",f"Enter {param}:",decimals=2)
                if not ok: return
                if value>0: params[param]=value; break
                QMessageBox.warning(self,"Error",f"{param} must be positive")
        self.items[name]=params
        save_to_excel(self.items,self.trucks)
        self.refresh_selection_tabs()

    def modify_item_dialog(self):
        if not self.items: return
        item,ok=QInputDialog.getItem(self,"Modify Item","Select item:",list(self.items.keys()),0,False)
        if not ok: return
        params=self.items[item].copy()
        for param in ["Length","Width","Height","weight"]:
            while True:
                value,ok=QInputDialog.getDouble(self,f"Modify {item}",f"Enter new {param}:",value=params[param],decimals=2)
                if not ok: return
                if value>0: params[param]=value; break
                QMessageBox.warning(self,"Error",f"{param} must be positive")
        self.items[item]=params
        save_to_excel(self.items,self.trucks)
        self.refresh_selection_tabs()

    def delete_item_dialog(self):
        if not self.items: return
        item,ok=QInputDialog.getItem(self,"Delete Item","Select item:",list(self.items.keys()),0,False)
        if not ok: return
        reply=QMessageBox.question(self,"Confirm Delete",f"Delete {item}?",QMessageBox.Yes|QMessageBox.No)
        if reply==QMessageBox.Yes:
            del self.items[item]
            save_to_excel(self.items,self.trucks)
            self.refresh_selection_tabs()

    def add_truck_dialog(self):
        name,ok=QInputDialog.getText(self,"New Truck","Enter truck name:")
        if not ok or not name.strip(): return
        name=name.strip()
        if name in self.trucks:
            QMessageBox.warning(self,"Error",f"Truck {name} exists"); return
        params={}
        for param in ["Length","Width","Height","max_weight"]:
            while True:
                value,ok=QInputDialog.getDouble(self,"New Truck",f"Enter {param}:",decimals=2)
                if not ok: return
                if value>0: params[param]=value; break
                QMessageBox.warning(self,"Error",f"{param} must be positive")
        self.trucks[name]=params
        save_to_excel(self.items,self.trucks)
        self.refresh_selection_tabs()

    def modify_truck_dialog(self):
        if not self.trucks: return
        truck,ok=QInputDialog.getItem(self,"Modify Truck","Select truck:",list(self.trucks.keys()),0,False)
        if not ok: return
        params=self.trucks[truck].copy()
        for param in ["Length","Width","Height","max_weight"]:
            while True:
                value,ok=QInputDialog.getDouble(self,f"Modify {truck}",f"Enter new {param}:",value=params[param],decimals=2)
                if not ok: return
                if value>0: params[param]=value; break
                QMessageBox.warning(self,"Error",f"{param} must be positive")
        self.trucks[truck]=params
        save_to_excel(self.items,self.trucks)
        self.refresh_selection_tabs()

    def delete_truck_dialog(self):
        if not self.trucks: return
        truck,ok=QInputDialog.getItem(self,"Delete Truck","Select truck:",list(self.trucks.keys()),0,False)
        if not ok: return
        reply=QMessageBox.question(self,"Confirm Delete",f"Delete {truck}?",QMessageBox.Yes|QMessageBox.No)
        if reply==QMessageBox.Yes:
            del self.trucks[truck]
            save_to_excel(self.items,self.trucks)
            self.refresh_selection_tabs()

    def refresh_selection_tabs(self):
        # Refresh item checkboxes
        for i,cb in self.item_checkboxes.items():
            cb.setText(f"{i} (Length:{self.items[i]['L']} Width:{self.items[i]['W']} Height:{self.items[i]['H']} Weight:{self.items[i]['weight']})")
        # Refresh truck checkboxes
        for t,cb in self.truck_checkboxes.items():
            cb.setText(f"{t} (Length:{self.trucks[t]['L']} Width:{self.trucks[t]['W']} Height:{self.trucks[t]['H']} MaxWeight:{self.trucks[t]['max_weight']})")

    # ------------------------
    # Optimization
    # ------------------------
    def run_optimization(self):
        if not self.selected_items or not self.selected_trucks:
            QMessageBox.warning(self,"Error","Select items and trucks first"); return
        item_qty={i:self.spin_boxes[i].value() for i in self.selected_items}
        trucks_subset={t:self.trucks[t] for t in self.selected_trucks}
        items_subset={i:self.items[i] for i in self.selected_items}
        self.solution,self.truck_weights=optimize_packing_with_items(items_subset,item_qty,trucks_subset)

        # populate table
        self.table.setRowCount(0)
        for t, items_list in self.solution.items():
            if items_list:
                total_weight=self.truck_weights[t]
                max_weight=self.trucks[t]["max_weight"]
                counts={it["item"]:0 for it in items_list}
                for it in items_list:
                    counts[it["item"]] +=1
                fullness = round(100*total_weight/max_weight,2)
                for item, count in counts.items():
                    row = self.table.rowCount()
                    self.table.insertRow(row)
                    self.table.setItem(row,0,QTableWidgetItem(t))
                    self.table.setItem(row,1,QTableWidgetItem(item))
                    self.table.setItem(row,2,QTableWidgetItem(str(count)))
                    self.table.setItem(row,3,QTableWidgetItem(str(count*self.items[item]["weight"])))
                    self.table.setItem(row,4,QTableWidgetItem(str(fullness)))
        QMessageBox.information(self,"Optimization","Optimization completed!")

    def save_solution_dialog(self):
        if not self.solution: return
        save_solution_to_excel(self.solution,self.truck_weights,self.items,self.trucks)
        QMessageBox.information(self,"Saved","Solution saved to Excel")
        self.load_history()  # refresh history tab

    def show_2d(self):
        if not self.solution: QMessageBox.warning(self,"Error","Run optimization first"); return
        self.window2d = Packing2DWindow(self.solution,self.trucks)
        self.window2d.show()

    # ------------------------
    # History
    # ------------------------
    def load_history(self):
        self.history_table.setRowCount(0)
        if not Path(EXCEL_FILE).exists(): 
            return
        wb = load_workbook(EXCEL_FILE, data_only=True)
        sheets = [s for s in wb.sheetnames if s.startswith("Solution_")]
        all_rows = []
        for s in sheets:
            ws = wb[s]
            solution_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                solution_rows.append([cell if cell is not None else "-" for cell in row])
            all_rows.append(solution_rows)  # keep each solution separate

        if not all_rows: 
            return

        n_cols = len(all_rows[0][0])
        self.history_table.setColumnCount(n_cols)
        self.history_table.setHorizontalHeaderLabels(["Truck","Weight","Item","Count","Item Weight","Fullness %"])

        for solution_rows in all_rows:
            for row_data in solution_rows:
                row = self.history_table.rowCount()
                self.history_table.insertRow(row)
                for col, val in enumerate(row_data):
                    self.history_table.setItem(row, col, QTableWidgetItem(str(val)))
            blank_row = self.history_table.rowCount()
            self.history_table.insertRow(blank_row)


    def delete_history(self):
        reply = QMessageBox.question(self,"Confirm","Delete all saved history?",QMessageBox.Yes|QMessageBox.No)
        if reply == QMessageBox.Yes:
            delete_all_history()
            QMessageBox.information(self,"Deleted","History deleted.")
            self.load_history()

# ------------------------
# MAIN
# ------------------------
if __name__=="__main__":
    app = QApplication(sys.argv)
    window = PackingWindow()
    window.show()
    sys.exit(app.exec_())
